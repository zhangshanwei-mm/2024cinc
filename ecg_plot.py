import os, sys, argparse
import numpy as np
import random
from scipy.io import savemat, loadmat
import matplotlib.pyplot as plt
import matplotlib
from matplotlib.ticker import AutoMinorLocator
from TemplateFiles.generate_template import generate_template
from helper_functions import convert_inches_to_seconds, convert_inches_to_volts, convert_mm_to_volts, \
    convert_mm_to_seconds
from math import ceil
from PIL import Image
from numpy import asarray
from random import randint
import matplotlib.patches as patches
import csv
import matplotlib.patches as patches

import cv2

from openpyxl import load_workbook


def newwrinkles(style_path, cont, tail):
    bimg = cv2.imread(style_path)
    bgray = cv2.cvtColor(bimg, cv2.COLOR_BGR2GRAY).astype(int)

    randnum = random.randint(1, 1000000)
    image_arr = np.array(cont)
    cv2.imwrite("./zhongjian" + tail + ".png", image_arr)
    fimg0 = cv2.imread("./zhongjian" + tail + ".png")
    fimg = cv2.resize(fimg0, dsize=bgray.shape, interpolation=cv2.INTER_NEAREST).astype(int)

    Y, X, N = fimg.shape
    exfimg = np.zeros((Y + 39, X + 39, N))
    exfimg[26:-13, 26:-13, :] = fimg  #
    for i in range(15):  #
        exfimg[26:-13, i, :] = fimg[:, 1, :]  #
        exfimg[26:-13, i + 13, :] = fimg[:, 1, :]  #
        exfimg[26:-13, -i - 1, :] = fimg[:, -1, :]  #
    for i in range(15):  #
        exfimg[i, :, :] = exfimg[28, :, :]  #
        exfimg[i + 13, :, :] = exfimg[28, :, :]  #
        exfimg[-i - 1, :, :] = exfimg[-14, :, :]  #

    goffset = (bgray - 128) / 10
    offsetLim1 = (np.floor(goffset) + 13).astype(int)
    offsetLim2 = (np.ceil(goffset) + 13).astype(int)
    sep1 = (goffset - np.floor(goffset)).flatten()
    sep2 = (np.ceil(goffset) - goffset).flatten()
    XX, YY = np.meshgrid(range(exfimg.shape[0] - 39), range(exfimg.shape[1] - 39))

    XX1, YY1 = XX + offsetLim1, YY + offsetLim1  #
    XX2, YY2 = XX + offsetLim2, YY + offsetLim2  #

    c1 = exfimg[YY1.flatten(), XX1.flatten(), :]
    c2 = exfimg[YY2.flatten(), XX2.flatten(), :]

    p1 = np.where(sep1 == 0, c1[:, 0], c2[:, 0] * sep1.flatten() + c1[:, 0] * sep2.flatten())
    p2 = np.where(sep1 == 0, c1[:, 1], c2[:, 1] * sep1.flatten() + c1[:, 1] * sep2.flatten())
    p3 = np.where(sep1 == 0, c1[:, 2], c2[:, 2] * sep1.flatten() + c1[:, 2] * sep2.flatten())

    newarr = np.array([p1.reshape(bgray.shape), p2.reshape(bgray.shape), p3.reshape(bgray.shape)])

    newarr = newarr.transpose((1, 2, 0)).astype(np.uint8)
    data = cv2.resize(newarr, dsize=(1100, 850))

    # cv2.imwrite("/data/0shared_data/ptb__ECGImage/ptbxl_data_new_change_with_log/newmask/masktest/filename1.png", data)
    return data


standard_values = {'y_grid_size': 0.5,
                   'x_grid_size': 0.2,
                   'y_grid_inch': 5 / 25.4,
                   'x_grid_inch': 5 / 25.4,
                   'grid_line_width': 0.5,
                   'lead_name_offset': 0.5,
                   'lead_fontsize': 11,
                   'x_gap': 1,
                   'y_gap': 0.5,
                   'display_factor': 1,
                   'line_width': 0.75,
                   'row_height': 8,
                   'dc_offset_length': 0.2,
                   'lead_length': 3,
                   'V1_length': 12,
                   'width': 11,
                   'height': 8.5
                   }

standard_major_colors = {'colour1': (0.4274, 0.196, 0.1843),  # brown
                         'colour2': (1, 0.796, 0.866),  # pink
                         'colour3': (0.0, 0.0, 0.4),  # blue
                         'colour4': (0, 0.3, 0.0),  # green
                         'colour5': (1, 0, 0)  # red
                         }

standard_minor_colors = {'colour1': (0.5882, 0.4196, 0.3960),
                         'colour2': (0.996, 0.9294, 0.9725),
                         'colour3': (0.0, 0, 0.7),
                         'colour4': (0, 0.8, 0.3),
                         'colour5': (0.996, 0.8745, 0.8588)
                         }

papersize_values = {'A0': (33.1, 46.8),
                    'A1': (33.1, 23.39),
                    'A2': (16.54, 23.39),
                    'A3': (11.69, 16.54),
                    'A4': (8.27, 11.69),
                    'letter': (8.5, 11)
                    }

leadNames_12 = ["III", 'aVF', 'V3', 'V6', 'II', 'aVL', 'V2', 'V5', 'I', 'aVR', 'V1', 'V4']


def inches_to_dots(value, resolution):
    return (value * resolution)


# Function to plot raw ecg signal
def ecg_plot(
        style_path,
        ecg,
        sample_rate,
        columns,
        rec_file_name,
        output_dir,
        resolution,
        pad_inches,
        lead_index,
        full_mode,
        store_text_bbox,
        full_header_file,
        units='',
        papersize='',
        x_gap=standard_values['x_gap'],
        y_gap=standard_values['y_gap'],
        display_factor=standard_values['display_factor'],
        line_width=standard_values['line_width'],
        title='',
        style=None,
        row_height=standard_values['row_height'],
        show_lead_name=True,
        show_grid=False,
        show_dc_pulse=False,
        y_grid=0,
        x_grid=0,
        standard_colours=False,
        bbox=False,
        print_txt=False,
        ful_mode=0
):
    # Inputs :
    # ecg - Dictionary of ecg signal with lead names as keys
    # sample_rate - Sampling rate of the ecg signal
    # lead_index - Order of lead indices to be plotted
    # columns - Number of columns to be plotted in each row
    # x_gap - gap between paper x axis border and signal plot
    # y_gap - gap between paper y axis border and signal plot
    # line_width - Width of line tracing the ecg
    # title - Title of figure
    # style - Black and white or colour
    # row_height - gap between corresponding ecg rows
    # show_lead_name - Option to show lead names or skip
    # show_dc_pulse - Option to show dc pulse
    # show_grid - Turn grid on or off

    # Initialize some params
    # secs represents how many seconds of ecg are plotted
    # leads represent number of leads in the ecg
    # rows are calculated based on corresponding number of leads and number of columns

    ifaddbara = random.randrange(1, 10)
    if ifaddbara > 5:
        ifaddbar = 1
    else:
        ifaddbar = 0

    matplotlib.use("Agg")
    randindex = randint(0, 99)
    random_sampler = random.uniform(-0.05, 0.004)

    # check if the ecg dict is empty
    if ecg == {}:
        return

    secs = len(list(ecg.items())[0][1]) / sample_rate

    leads = len(lead_index)

    rows = int(ceil(leads / columns))

    if (full_mode != 'None'):
        rows += 1
        leads += 1

    # Grid calibration
    # Each big grid corresponds to 0.2 seconds and 0.5 mV
    # To do: Select grid size in a better way
    y_grid_size = standard_values['y_grid_size']
    x_grid_size = standard_values['x_grid_size']
    grid_line_width = standard_values['grid_line_width']
    lead_name_offset = standard_values['lead_name_offset']
    lead_fontsize = standard_values['lead_fontsize']

    # Set max and min coordinates to mark grid. Offset x_max slightly (i.e by 1 column width)

    if papersize == '':
        width = standard_values['width']
        height = standard_values['height']
    else:
        width = papersize_values[papersize][1]
        height = papersize_values[papersize][0]

    y_grid = standard_values['y_grid_inch']
    x_grid = standard_values['x_grid_inch']
    y_grid_dots = y_grid * resolution
    x_grid_dots = x_grid * resolution

    # row_height = height * y_grid_size/(y_grid*(rows+2))
    row_height = (height * y_grid_size / y_grid) / (rows + 2)
    x_max = width * x_grid_size / x_grid
    x_min = 0
    x_gap = np.floor(((x_max - (columns * secs)) / 2) / 0.2) * 0.2
    y_min = 0
    y_max = height * y_grid_size / y_grid

    # Set figure and subplot sizes
    fig, ax = plt.subplots(figsize=(width, height))

    fig.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig.suptitle(title)

    # Mark grid based on whether we want black and white or colour
    fig1, ax1 = plt.subplots(figsize=(width, height))
    fig2, ax2 = plt.subplots(figsize=(width, height))
    fig3, ax3 = plt.subplots(figsize=(width, height))
    fig4, ax4 = plt.subplots(figsize=(width, height))
    fig5, ax5 = plt.subplots(figsize=(width, height))
    fig6, ax6 = plt.subplots(figsize=(width, height))
    fig7, ax7 = plt.subplots(figsize=(width, height))
    fig8, ax8 = plt.subplots(figsize=(width, height))
    fig9, ax9 = plt.subplots(figsize=(width, height))
    fig10, ax10 = plt.subplots(figsize=(width, height))
    fig11, ax11 = plt.subplots(figsize=(width, height))
    fig12, ax12 = plt.subplots(figsize=(width, height))
    fig13, ax13 = plt.subplots(figsize=(width, height))
    fig14, ax14 = plt.subplots(figsize=(width, height))

    fig1.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig1.suptitle(title)

    fig2.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig2.suptitle(title)

    fig3.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig3.suptitle(title)

    fig4.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig4.suptitle(title)

    fig5.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig5.suptitle(title)

    fig6.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig6.suptitle(title)

    fig7.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig7.suptitle(title)

    fig8.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig8.suptitle(title)

    fig9.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig9.suptitle(title)

    fig10.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig10.suptitle(title)

    fig11.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig11.suptitle(title)

    fig12.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig12.suptitle(title)

    fig13.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig13.suptitle(title)

    fig14.subplots_adjust(
        hspace=0,
        wspace=0,
        left=0,
        right=1,
        bottom=0,
        top=1
    )

    fig14.suptitle(title)

    if (style == 'bw'):
        color_major = (0.4, 0.4, 0.4)
        color_minor = (0.75, 0.75, 0.75)
        color_line = (0, 0, 0)
    elif (standard_colours > 0):
        random_colour_index = standard_colours
        color_major = standard_major_colors['colour' + str(random_colour_index)]
        color_minor = standard_minor_colors['colour' + str(random_colour_index)]
        randcolorindex_grey = randint(0, 24)
        grey_random_color = random.uniform(0, 0.2)
        color_line = (grey_random_color, grey_random_color, grey_random_color)
    else:
        randcolorindex_red = randint(0, 24)
        major_random_color_sampler_red = random.uniform(0, 0.8)
        randcolorindex_green = randint(0, 24)
        major_random_color_sampler_green = random.uniform(0, 0.5)
        randcolorindex_blue = randint(0, 24)
        major_random_color_sampler_blue = random.uniform(0, 0.5)

        randcolorindex_minor = randint(0, 24)
        minor_offset = random.uniform(0, 0.2)
        minor_random_color_sampler_red = major_random_color_sampler_red + minor_offset
        minor_random_color_sampler_green = random.uniform(0, 0.5) + minor_offset
        minor_random_color_sampler_blue = random.uniform(0, 0.5) + minor_offset

        randcolorindex_grey = randint(0, 24)
        grey_random_color = random.uniform(0, 0.2)
        color_major = (
        major_random_color_sampler_red, major_random_color_sampler_green, major_random_color_sampler_blue)
        color_minor = (
        minor_random_color_sampler_red, minor_random_color_sampler_green, minor_random_color_sampler_blue)

        color_line = (grey_random_color, grey_random_color, grey_random_color)

    # Set grid
    # Standard ecg has grid size of 0.5 mV and 0.2 seconds. Set ticks accordingly
    if (show_grid):
        ax.set_xticks(np.arange(x_min, x_max, x_grid_size))
        ax.set_yticks(np.arange(y_min, y_max, y_grid_size))
        ax.minorticks_on()

        ax.xaxis.set_minor_locator(AutoMinorLocator(5))

        # set grid line style
        ax.grid(which='major', linestyle='-', linewidth=grid_line_width, color=color_major)

        ax.grid(which='minor', linestyle='-', linewidth=grid_line_width, color=color_minor)

    else:
        ax.grid(False)
    ax.set_ylim(y_min, y_max)
    ax.set_xlim(x_min, x_max)
    ax.tick_params(axis='x', colors='white')
    ax.tick_params(axis='y', colors='white')
    ax1.set_ylim(y_min, y_max)
    ax1.set_xlim(x_min, x_max)
    ax1.tick_params(axis='x', colors='white')
    ax1.tick_params(axis='y', colors='white')

    ax2.set_ylim(y_min, y_max)
    ax2.set_xlim(x_min, x_max)
    ax2.tick_params(axis='x', colors='white')
    ax2.tick_params(axis='y', colors='white')

    ax3.set_ylim(y_min, y_max)
    ax3.set_xlim(x_min, x_max)
    ax3.tick_params(axis='x', colors='white')
    ax3.tick_params(axis='y', colors='white')

    ax4.set_ylim(y_min, y_max)
    ax4.set_xlim(x_min, x_max)
    ax4.tick_params(axis='x', colors='white')
    ax4.tick_params(axis='y', colors='white')

    ax5.set_ylim(y_min, y_max)
    ax5.set_xlim(x_min, x_max)
    ax5.tick_params(axis='x', colors='white')
    ax5.tick_params(axis='y', colors='white')

    ax6.set_ylim(y_min, y_max)
    ax6.set_xlim(x_min, x_max)
    ax6.tick_params(axis='x', colors='white')
    ax6.tick_params(axis='y', colors='white')

    ax7.set_ylim(y_min, y_max)
    ax7.set_xlim(x_min, x_max)
    ax7.tick_params(axis='x', colors='white')
    ax7.tick_params(axis='y', colors='white')

    ax8.set_ylim(y_min, y_max)
    ax8.set_xlim(x_min, x_max)
    ax8.tick_params(axis='x', colors='white')
    ax8.tick_params(axis='y', colors='white')

    ax9.set_ylim(y_min, y_max)
    ax9.set_xlim(x_min, x_max)
    ax9.tick_params(axis='x', colors='white')
    ax9.tick_params(axis='y', colors='white')

    ax10.set_ylim(y_min, y_max)
    ax10.set_xlim(x_min, x_max)
    ax10.tick_params(axis='x', colors='white')
    ax10.tick_params(axis='y', colors='white')

    ax11.set_ylim(y_min, y_max)
    ax11.set_xlim(x_min, x_max)
    ax11.tick_params(axis='x', colors='white')
    ax11.tick_params(axis='y', colors='white')

    ax12.set_ylim(y_min, y_max)
    ax12.set_xlim(x_min, x_max)
    ax12.tick_params(axis='x', colors='white')
    ax12.tick_params(axis='y', colors='white')

    ax13.set_ylim(y_min, y_max)
    ax13.set_xlim(x_min, x_max)
    ax13.tick_params(axis='x', colors='white')
    ax13.tick_params(axis='y', colors='white')

    ax14.set_ylim(y_min, y_max)
    ax14.set_xlim(x_min, x_max)
    ax14.tick_params(axis='x', colors='white')
    ax14.tick_params(axis='y', colors='white')

    tall = []
    axll = []
    tall.append(fig1)
    tall.append(fig2)
    tall.append(fig3)
    tall.append(fig4)
    tall.append(fig5)
    tall.append(fig6)
    tall.append(fig7)
    tall.append(fig8)
    tall.append(fig9)
    tall.append(fig10)
    tall.append(fig11)
    tall.append(fig12)
    tall.append(fig13)

    axll.append(ax1)
    axll.append(ax2)
    axll.append(ax3)
    axll.append(ax4)
    axll.append(ax5)
    axll.append(ax6)
    axll.append(ax7)
    axll.append(ax8)
    axll.append(ax9)
    axll.append(ax10)
    axll.append(ax11)
    axll.append(ax12)
    axll.append(ax13)

    lead_num = 0

    # Step size will be number of seconds per sample i.e 1/sampling_rate
    step = (1.0 / sample_rate)

    dc_offset = 0
    standard_values['dc_offset_length'] = 0.1
    if (show_dc_pulse):
        dc_offset = sample_rate * standard_values['dc_offset_length'] * step

    # Iterate through each lead in lead_index array.
    y_offset = (row_height / 2)
    x_offset = 0

    text_bbox = []
    lead_bbox = []

    if columns == 1:
        leadNames_12 = ['V6', 'V5', 'V4', 'V3', 'V2', 'V1', 'aVF', 'aVL', 'aVR', 'III', "II", 'I']
        deviation_x = 0.3
        deviation_y = 1.2

    if columns == 2:
        leadNames_12 = ['aVF', 'V6', 'aVL', 'V5', 'aVR', 'V4', 'III', 'V3', "II", 'V2', 'I', 'V1']
        deviation_x = 0.3
        deviation_y = 1.2

    if columns == 4:
        leadNames_12 = ["III", 'aVF', 'V3', 'V6', 'II', 'aVL', 'V2', 'V5', 'I', 'aVR', 'V1', 'V4']
        deviation_x = 0.3
        deviation_y = 1.2

    mask = np.full((13, 1100, 850), False)
    maskw = np.full((13, 1100, 850), False)
    for i in np.arange(len(lead_index)):

        if len(lead_index) == 12:
            leadName = leadNames_12[i]
        else:
            leadName = lead_index[i]

        # y_offset is computed by shifting by a certain offset based on i, and also by row_height/2 to account for half the waveform below the axis
        if (i % columns == 0):
            y_offset += row_height

        # x_offset will be distance by which we shift the plot in each iteration
        if (columns > 1):
            x_offset = (i % columns) * secs

        else:
            x_offset = 0

        # Create dc pulse wave to plot at the beginning of plot. Dc pulse will be 0.2 seconds
        x_range = np.arange(0, sample_rate * standard_values['dc_offset_length'] * step + 4 * step, step)
        dc_pulse = np.ones(len(x_range))
        dc_pulse = np.concatenate(((0, 0), dc_pulse[2:-2], (0, 0)))

        # Print lead name at .5 ( or 5 mm distance) from plot
        if (show_lead_name):

            t1 = ax.text(x_offset + x_gap + deviation_x,
                         y_offset - lead_name_offset - 0.2 + deviation_y,
                         leadName,
                         fontsize=lead_fontsize)

            if (store_text_bbox):
                renderer1 = fig.canvas.get_renderer()
                transf = ax.transData.inverted()
                bb = t1.get_window_extent()
                x1 = bb.x0 * resolution / fig.dpi
                y1 = bb.y0 * resolution / fig.dpi
                x2 = bb.x1 * resolution / fig.dpi
                y2 = bb.y1 * resolution / fig.dpi
                text_bbox.append([x1, y1, x2, y2, leadName])

        # If we are plotting the first row-1 plots, we plot the dc pulse prior to adding the waveform
        if (columns == 1 and i in np.arange(0, rows)):

            if (show_dc_pulse):
                # Plot dc pulse for 0.2 seconds with 2 trailing and leading zeros to get the pulse
                t1 = ax.plot(x_range + x_offset + x_gap,
                             dc_pulse + y_offset,
                             linewidth=line_width * 1.5,
                             color=color_line
                             )
                if (bbox):
                    renderer1 = fig.canvas.get_renderer()
                    transf = ax.transData.inverted()
                    bb = t1[0].get_window_extent()
                    x1, y1 = bb.x0 * resolution / fig.dpi, bb.y0 * resolution / fig.dpi
                    x2, y2 = bb.x1 * resolution / fig.dpi, bb.y1 * resolution / fig.dpi


        elif (columns == 2 and (i == 10 or i == 8 or i == 6 or i == 4 or i == 2 or i == 0)):

            if (show_dc_pulse):
                # Plot dc pulse for 0.2 seconds with 2 trailing and leading zeros to get the pulse
                t1 = ax.plot(np.arange(0, sample_rate * standard_values['dc_offset_length'] * step + 4 * step,
                                       step) + x_offset + x_gap,
                             dc_pulse + y_offset,
                             linewidth=line_width * 1.5,
                             color=color_line
                             )
                if (bbox):
                    renderer1 = fig.canvas.get_renderer()
                    transf = ax.transData.inverted()
                    bb = t1[0].get_window_extent()
                    x1, y1 = bb.x0 * resolution / fig.dpi, bb.y0 * resolution / fig.dpi
                    x2, y2 = bb.x1 * resolution / fig.dpi, bb.y1 * resolution / fig.dpi




        elif (columns == 4 and (i == 0 or i == 4 or i == 8)):

            if (show_dc_pulse):
                # Plot dc pulse for 0.2 seconds with 2 trailing and leading zeros to get the pulse
                t1 = ax.plot(np.arange(0, sample_rate * standard_values['dc_offset_length'] * step + 4 * step,
                                       step) + x_offset + x_gap,
                             dc_pulse + y_offset,
                             linewidth=line_width * 1.5,
                             color=color_line
                             )
                if (bbox):
                    renderer1 = fig.canvas.get_renderer()
                    transf = ax.transData.inverted()
                    bb = t1[0].get_window_extent()
                    x1, y1 = bb.x0 * resolution / fig.dpi, bb.y0 * resolution / fig.dpi
                    x2, y2 = bb.x1 * resolution / fig.dpi, bb.y1 * resolution / fig.dpi

        if (show_dc_pulse):
            y_0 = ecg[leadName][0]
            yc = y_offset - y_0
            xc = 0.05
        else:
            yc = y_offset
            xc = 0

        t1 = ax.plot(np.arange(0, len(ecg[leadName]) * step, step) + x_offset + dc_offset + x_gap + xc,
                     ecg[leadName] + yc,
                     linewidth=line_width,
                     color=color_line
                     )

        t2 = ax14.plot(np.arange(0, len(ecg[leadName]) * step, step) + x_offset + dc_offset + x_gap + xc,
                       # ecg[leadName] + y_offset,
                       ecg[leadName] + yc,
                       linewidth=line_width,
                       color=color_line
                       )

        if columns == 1 and i == 0:
            jj = 11
        if columns == 1 and i == 1:
            jj = 10
        if columns == 1 and i == 2:
            jj = 9
        if columns == 1 and i == 3:
            jj = 8
        if columns == 1 and i == 4:
            jj = 7
        if columns == 1 and i == 5:
            jj = 6
        if columns == 1 and i == 6:
            jj = 5
        if columns == 1 and i == 7:
            jj = 4
        if columns == 1 and i == 8:
            jj = 3
        if columns == 1 and i == 9:
            jj = 2
        if columns == 1 and i == 10:
            jj = 1
        if columns == 1 and i == 11:
            jj = 0

        if columns == 2 and i == 0:
            jj = 5
        if columns == 2 and i == 1:
            jj = 11
        if columns == 2 and i == 2:
            jj = 4
        if columns == 2 and i == 3:
            jj = 10
        if columns == 2 and i == 4:
            jj = 3
        if columns == 2 and i == 5:
            jj = 9
        if columns == 2 and i == 6:
            jj = 2
        if columns == 2 and i == 7:
            jj = 8
        if columns == 2 and i == 8:
            jj = 1
        if columns == 2 and i == 9:
            jj = 7
        if columns == 2 and i == 10:
            jj = 0
        if columns == 2 and i == 11:
            jj = 6

        if columns == 4 and i == 0:
            jj = 2
        if columns == 4 and i == 1:
            jj = 5
        if columns == 4 and i == 2:
            jj = 8
        if columns == 4 and i == 3:
            jj = 11
        if columns == 4 and i == 4:
            jj = 1
        if columns == 4 and i == 5:
            jj = 4
        if columns == 4 and i == 6:
            jj = 7
        if columns == 4 and i == 7:
            jj = 10
        if columns == 4 and i == 8:
            jj = 0
        if columns == 4 and i == 9:
            jj = 3
        if columns == 4 and i == 10:
            jj = 6
        if columns == 4 and i == 11:
            jj = 9

        if (1 == 1):
            renderer1 = fig.canvas.get_renderer()
            transf = ax.transData.inverted()
            bb = t1[0].get_window_extent()
            if show_dc_pulse == False or (columns == 4 and (i != 0 and i != 4 and i != 8)):
                x1, y1 = bb.x0 * resolution / fig.dpi, bb.y0 * resolution / fig.dpi
                x2, y2 = bb.x1 * resolution / fig.dpi, bb.y1 * resolution / fig.dpi
            else:
                y1 = min(y1, bb.y0 * resolution / fig.dpi)
                y2 = max(y2, bb.y1 * resolution / fig.dpi)
                x2 = bb.x1 * resolution / fig.dpi

            lead_bbox.append([x1, y1, x2, y2, leadName])

        t = axll[i].plot(np.arange(0, len(ecg[leadName]) * step, step) + x_offset + dc_offset + x_gap + xc,
                         # ecg[leadName] + y_offset,
                         ecg[leadName] + yc,
                         linewidth=line_width,
                         color=color_line
                         )
        tall[i].canvas.draw()
        # tall[i].savefig("/data/0shared_data/ptb__ECGImage/ptbxl_data_new_change_with_log/test1000/" + tail + str(i) + ".png",dpi=100)
        fig_str = tall[i].canvas.tostring_rgb()
        data = np.frombuffer(fig_str, dtype=np.uint8).reshape(850, 1100, 3)
        image_arr = np.array(data)

        asd = 0
        if x1 == float('inf') or y1 == float('inf') or x2 == float('inf') or y2 == float('inf'):
            asd = 1
        if x1 == float('-inf') or y1 == float('-inf') or x2 == float('-inf') or y2 == float('-inf'):
            asd = 1

        if asd == 0:
            x11 = 0
            y11 = 0
            x12 = 1099
            y12 = 849

            if int(x1 / 2) - 6 > 0:
                x11 = int(x1 / 2) - 6
            if int(x2 / 2) + 6 < 1099:
                x12 = int(x2 / 2) + 6
            if int(float(850 - y2 / 2) - 1) - 6 > 0:
                y11 = int(float(850 - y2 / 2) - 1) - 6
            if int(float(850 - y1 / 2) - 1) + 6 < 849:
                y12 = int(float(850 - y1 / 2) - 1) + 6
            for x in range(x11, x12):
                for y in range(y11, y12):
                    if image_arr[y][x][0] != 255 or image_arr[y][x][1] != 255 or image_arr[y][x][2] != 255:
                        mask[jj][x][y] = True
        head, tail = os.path.split(rec_file_name)
        image_arr = newwrinkles(style_path, image_arr, tail)
        asd = 0
        if x1 == float('inf') or y1 == float('inf') or x2 == float('inf') or y2 == float('inf'):
            asd = 1
        if x1 == float('-inf') or y1 == float('-inf') or x2 == float('-inf') or y2 == float('-inf'):
            asd = 1

        if asd == 0:
            x11 = 0
            y11 = 0
            x12 = 1099
            y12 = 849

            if int(x1 / 2) - 6 > 0:
                x11 = int(x1 / 2) - 6
            if int(x2 / 2) + 6 < 1099:
                x12 = int(x2 / 2) + 6
            if int(float(850 - y2 / 2) - 1) - 6 > 0:
                y11 = int(float(850 - y2 / 2) - 1) - 6
            if int(float(850 - y1 / 2) - 1) + 6 < 849:
                y12 = int(float(850 - y1 / 2) - 1) + 6
            for x in range(x11, x12):
                for y in range(y11, y12):
                    if image_arr[y][x][0] != 255 or image_arr[y][x][1] != 255 or image_arr[y][x][2] != 255:
                        maskw[jj][x][y] = True

        start_ind = round((x_offset + dc_offset + x_gap) * x_grid_dots / x_grid_size)
        end_ind = round((x_offset + dc_offset + x_gap + len(ecg[leadName]) * step) * x_grid_dots / x_grid_size)
        tickSize_step = 0.002
        tickLength = 8
        if ifaddbar == 1:
            if columns > 1 and (i + 1) % columns != 0:
                sep_x = [len(ecg[leadName]) * step + x_offset + dc_offset + x_gap] * round(8 * y_grid_dots)
                sep_x = np.array(sep_x)
                sep_y = np.linspace(y_offset - tickLength / 2 * y_grid_dots * tickSize_step,
                                    y_offset + tickSize_step * y_grid_dots * tickLength / 2, len(sep_x))
                ax.plot(sep_x, sep_y, linewidth=line_width * 3, color=color_line)

    # Plotting longest lead for 12 seconds

    if ful_mode == 1:
        a = 1
    elif (full_mode != 'None'):
        if (show_lead_name):
            t1 = ax.text(x_gap + deviation_x,
                         row_height / 2 - lead_name_offset + deviation_y,
                         full_mode,
                         fontsize=lead_fontsize)

            if (store_text_bbox):
                renderer1 = fig.canvas.get_renderer()
                transf = ax.transData.inverted()
                bb = t1.get_window_extent(renderer=fig.canvas.renderer)
                x1 = bb.x0 * resolution / fig.dpi
                y1 = bb.y0 * resolution / fig.dpi
                x2 = bb.x1 * resolution / fig.dpi
                y2 = bb.y1 * resolution / fig.dpi
                text_bbox.append([x1, y1, x2, y2, full_mode])

        if (show_dc_pulse):
            t1 = ax.plot(x_range + x_gap,
                         dc_pulse + row_height / 2 - lead_name_offset + 0.8,
                         linewidth=line_width * 1.5,
                         color=color_line
                         )

            if (bbox):
                renderer1 = fig.canvas.get_renderer()
                transf = ax.transData.inverted()
                bb = t1[0].get_window_extent()
                x1, y1 = bb.x0 * resolution / fig.dpi, bb.y0 * resolution / fig.dpi
                x2, y2 = bb.x1 * resolution / fig.dpi, bb.y1 * resolution / fig.dpi

        dc_full_lead_offset = 0
        if (show_dc_pulse):
            dc_full_lead_offset = sample_rate * standard_values['dc_offset_length'] * step

        if (show_dc_pulse):
            y_0 = ecg['full' + full_mode][0]
            xc = 0.05
        else:
            y_0 = 0
            xc = 0

        t1 = ax.plot(np.arange(0, len(ecg['full' + full_mode]) * step, step) + x_gap + dc_full_lead_offset + xc,
                     ecg['full' + full_mode] + row_height / 2 - lead_name_offset + 0.8 - y_0,
                     linewidth=line_width,
                     color=color_line
                     )

        t2 = ax14.plot(np.arange(0, len(ecg[leadName]) * step, step) + x_offset + dc_offset + x_gap + xc,
                       # ecg[leadName] + y_offset,
                       ecg[leadName] + yc,
                       linewidth=line_width,
                       color=color_line
                       )

        if (1 == 1):
            renderer1 = fig.canvas.get_renderer()
            transf = ax.transData.inverted()
            bb = t1[0].get_window_extent()
            if show_dc_pulse == False:
                x1, y1 = bb.x0 * resolution / fig.dpi, bb.y0 * resolution / fig.dpi
                x2, y2 = bb.x1 * resolution / fig.dpi, bb.y1 * resolution / fig.dpi
            else:
                y1 = min(y1, bb.y0 * resolution / fig.dpi)
                y2 = max(y2, bb.y1 * resolution / fig.dpi)
                x2 = bb.x1 * resolution / fig.dpi

        t = axll[12].plot(np.arange(0, len(ecg['full' + full_mode]) * step, step) + x_gap + dc_full_lead_offset + xc,
                          ecg['full' + full_mode] + row_height / 2 - lead_name_offset + 0.8 - y_0,
                          linewidth=line_width,
                          color=color_line
                          )

        tall[12].canvas.draw()
        fig_str = tall[12].canvas.tostring_rgb()
        data = np.frombuffer(fig_str, dtype=np.uint8).reshape(850, 1100, 3)
        image_arr = np.array(data)

        asd = 0
        if x1 == float('inf') or y1 == float('inf') or x2 == float('inf') or y2 == float('inf'):
            asd = 1
        if x1 == float('-inf') or y1 == float('-inf') or x2 == float('-inf') or y2 == float('-inf'):
            asd = 1

        if asd == 0:
            x11 = 0
            y11 = 0
            x12 = 1099
            y12 = 849

            if int(x1 / 2) - 6 > 0:
                x11 = int(x1 / 2) - 6
            if int(x2 / 2) + 6 < 1099:
                x12 = int(x2 / 2) + 6
            if int(float(850 - y2 / 2) - 1) - 6 > 0:
                y11 = int(float(850 - y2 / 2) - 1) - 6
            if int(float(850 - y1 / 2) - 1) + 6 < 849:
                y12 = int(float(850 - y1 / 2) - 1) + 6
            for x in range(x11, x12):
                for y in range(y11, y12):
                    if image_arr[y][x][0] != 255 or image_arr[y][x][1] != 255 or image_arr[y][x][2] != 255:
                        mask[12][x][y] = True
        head, tail = os.path.split(rec_file_name)
        image_arr = newwrinkles(style_path, image_arr, tail)
        asd = 0
        if x1 == float('inf') or y1 == float('inf') or x2 == float('inf') or y2 == float('inf'):
            asd = 1
        if x1 == float('-inf') or y1 == float('-inf') or x2 == float('-inf') or y2 == float('-inf'):
            asd = 1

        if asd == 0:
            x11 = 0
            y11 = 0
            x12 = 1099
            y12 = 849

            if int(x1 / 2) - 6 > 0:
                x11 = int(x1 / 2) - 6
            if int(x2 / 2) + 6 < 1099:
                x12 = int(x2 / 2) + 6
            if int(float(850 - y2 / 2) - 1) - 6 > 0:
                y11 = int(float(850 - y2 / 2) - 1) - 6
            if int(float(850 - y1 / 2) - 1) + 6 < 849:
                y12 = int(float(850 - y1 / 2) - 1) + 6
            for x in range(x11, x12):
                for y in range(y11, y12):
                    if image_arr[y][x][0] != 255 or image_arr[y][x][1] != 255 or image_arr[y][x][2] != 255:
                        maskw[12][x][y] = True

            lead_bbox.append([x1, y1, x2, y2, 1])

        start_ind = round((dc_full_lead_offset + x_gap) * x_grid_dots / x_grid_size)
        end_ind = round((dc_full_lead_offset + x_gap + len(ecg['full' + full_mode]) * step) * x_grid_dots / x_grid_size)

    head, tail = os.path.split(rec_file_name)
    rec_file_name = os.path.join(output_dir, tail)

    # printed template file
    if print_txt:
        x_offset = 0.05
        y_offset = int(y_max)
        template_name = 'custom_template.png'
        printed_text, attributes, flag = generate_template(full_header_file)

        if flag:
            for l in range(0, len(printed_text), 1):

                for j in printed_text[l]:
                    curr_l = ''
                    if j in attributes.keys():
                        curr_l += str(attributes[j])
                    ax.text(x_offset, y_offset, curr_l, fontsize=lead_fontsize)
                    x_offset += 3

                y_offset -= 0.5
                x_offset = 0.05
        else:
            for line in printed_text:
                ax.text(x_offset, y_offset, line, fontsize=lead_fontsize)
                y_offset -= 0.5

    # change x and y res
    ax.text(2, 0.5, '25mm/s', fontsize=lead_fontsize)
    ax.text(4, 0.5, '10mm/mV', fontsize=lead_fontsize)

    fig.savefig(os.path.join(output_dir, tail + '.png'), dpi=100)

    fig14.savefig(os.path.join(output_dir, tail + 'white.png'), dpi=100)

    np.savez_compressed(os.path.join(output_dir, tail + '.npz'), mask=mask)
    np.savez_compressed(os.path.join(output_dir, tail + 'wrinkles.npz'), mask=maskw)

    plt.close(fig)
    plt.clf()
    plt.cla()

    if pad_inches != 0:
        ecg_image = Image.open(os.path.join(output_dir, tail + '.png'))

        right = pad_inches * resolution
        left = pad_inches * resolution
        top = pad_inches * resolution
        bottom = pad_inches * resolution
        width, height = ecg_image.size
        new_width = width + right + left
        new_height = height + top + bottom
        result_image = Image.new(ecg_image.mode, (new_width, new_height), (255, 255, 255))
        result_image.paste(ecg_image, (left, top))

        result_image.save(os.path.join(output_dir, tail + '.png'))

        plt.close('all')
        plt.close(fig)
        plt.clf()
        plt.cla()

    if (store_text_bbox):
        if (os.path.exists(os.path.join(output_dir, 'text_bounding_box')) == False):
            os.mkdir(os.path.join(output_dir, 'text_bounding_box'))

        with open(os.path.join(output_dir, 'text_bounding_box', tail + '.txt'), 'w') as f:
            for i, l in enumerate(text_bbox):
                if pad_inches != 0:
                    l[0] += left
                    l[2] += left
                    l[1] += top
                    l[3] += top

                for val in l[:4]:
                    f.write(str(val))
                    f.write(',')
                f.write(str(l[4]))
                f.write('\n')

    if (bbox):
        if (os.path.exists(os.path.join(output_dir, 'lead_bounding_box')) == False):
            os.mkdir(os.path.join(output_dir, 'lead_bounding_box'))
        with open(os.path.join(output_dir, 'lead_bounding_box', tail + '.txt'), 'w') as f:
            for i, l in enumerate(lead_bbox):
                if pad_inches != 0:
                    l[0] += left
                    l[2] += left
                    l[1] += top
                    l[3] += top

                for val in l[:4]:
                    f.write(str(val))
                    f.write(',')
                f.write(str(l[4]))
                f.write('\n')

    wb = load_workbook(output_dir + "log.xlsx")
    ws = wb['Sheet']
    row = ws.max_row + 1
    ws.cell(row, 1).value = tail
    ws.cell(row, 2).value = show_dc_pulse
    ws.cell(row, 3).value = columns
    ws.cell(row, 4).value = print_txt
    ws.cell(row, 5).value = show_lead_name
    wb.save(output_dir + "log.xlsx")

    return x_grid_dots, y_grid_dots
